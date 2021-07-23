from django.shortcuts import render
import openpyxl

def index(request):
    if "GET" == request.method:
        return render(request, 'xl_calculation/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file)

        # get all sheets
        sheets = wb.sheetnames
        print(sheets)

        # get a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        #getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        #reading a cell
        print(worksheet["A1"].value)

        #i terating over the rows and getting
        # value from each cell in row
        excel_data = list()

        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        return render(request, "xl_calculation/index.html", {"excel_data":excel_data})
