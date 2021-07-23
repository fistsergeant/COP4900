from django.shortcuts import render
from django.conf import settings
from django.http import HttpResponse, Http404
from django.shortcuts import render
import openpyxl
import os
from xl_calculation.staffNoWait import computeNoWait

fn = ''

def download(request):
    global fn
    #Only run if the file was already upload
    if fn == '':
        raise Http404

    with open(fn, 'rb') as fh:
        response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename={os.path.basename(fn)}' # name it after the original file title
        fn = ''
        return response
    

def index(request):
    global fn
    if "GET" == request.method:
        return render(request, 'xl_calculation/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        print(excel_file)
        wb = openpyxl.load_workbook(excel_file)
        
        print(type(str(excel_file)))
        # get all sheets
        sheets = wb.sheetnames
        print(sheets)

        # get a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        #getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        #reading a cell
        print(worksheet["A1"].value)

        #   numClassrooms = Number of classrooms                                       
        #   totalAtRiskStu = Total number of at risk students                            
        #   numStuPerTeacherOrAide = T2 Lite: Number of students per teacher or aide           
        #   probRespond = T2 Lite: Probability respond to T2 Lite                      
        #   numStuPerPsychologist = Evaluation: Number of students per psychologist            
        #   numStuPerInterventionist = T2 Heavy: Number of students per interventionist  

        numClassrooms = int(worksheet["F7"].value) 
        totalAtRiskStu = int(worksheet["F8"].value) 
        numStuPerTeacherOrAide = int(worksheet["F9"].value) 
        probRespond = float(worksheet["F10"].value) 
        numStuPerPsychologist = int(worksheet["F11"].value) 
        numStuPerInterventionist = int(worksheet["F12"].value) 

        # Using the function compuNoWait of the staffNoWait.py file
        numAides, numPsychologists, numInterventionists =                            \
        computeNoWait(numClassrooms, totalAtRiskStu, numStuPerTeacherOrAide,         \
            probRespond, numStuPerPsychologist, numStuPerInterventionist)

        # Placing the answer in the space of the answer
        worksheet['F15'] = numAides
        worksheet['F16'] = numPsychologists
        worksheet['F17'] = numInterventionists
        
        # Get file name of the file upload
        basename = os.path.basename(str(excel_file))
        #Add the complete file to the path files 
        fn = "files/" +basename 
        print(fn)
        # save file xlsx with the solution"
        wb.save(fn)

        excel_data = list()

        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                if cell.value == None: # if the cell is empty so replace it to a whitespace string
                    cell.value = ""
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)


        return render(request, "xl_calculation/index.html", {"excel_data":excel_data,"created":True})
